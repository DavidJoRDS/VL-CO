import streamlit as st
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
import requests
import os
import zipfile
import shutil
import re
from io import BytesIO
from PIL import Image as PILImage
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from urllib.parse import urljoin
from openpyxl.styles import Alignment, Font, PatternFill
from concurrent.futures import ThreadPoolExecutor, as_completed
import datetime

# ─────────────────────────────────────────
# 페이지 설정
# ─────────────────────────────────────────
st.set_page_config(page_title="VL&CO 상품크롤러", layout="wide")
st.title("🛒 VL&CO 상품크롤러")
st.caption("국내외 모든 쇼핑몰 범용 | 세일가 정밀 판별 | 빠른 병렬 수집")

# ─────────────────────────────────────────
# 세션 상태 초기화
# ─────────────────────────────────────────
for key in ['excel_data', 'zip_data', 'result_count']:
    if key not in st.session_state:
        st.session_state[key] = None
if 'logs' not in st.session_state:
    st.session_state.logs = []

target_url = st.text_input(
    "크롤링할 사이트 주소를 입력하세요",
    value="",
    placeholder="https://example.com/products"
)

# ─────────────────────────────────────────
# 유틸: 가격 파싱
# ─────────────────────────────────────────
PRICE_MARKERS = ['원', '₩', 'KRW', '$', '¥', '€', '£']

def parse_price_from_line(line: str) -> float:
    """
    한 줄 텍스트에서 가격 숫자 하나를 추출.
    - '%' 앞 숫자(할인율) 제거: '20% 143,200원' → 143200
    - '(숫자)' 리뷰수/괄호 제거: '(140)' → 제거
    - 쉼표 있는 숫자 우선, 그 다음 5자리+, 그 다음 화폐기호+숫자
    """
    s = re.sub(r'\b\d{1,3}\s*%', '', line)        # 할인율 제거 (20%, 30% 등)
    s = re.sub(r'\(\s*\d+\s*\)', '', s).strip()    # (140) 리뷰수 제거

    # 쉼표 있는 숫자 (1,000 이상) — 가장 신뢰도 높음
    m = re.search(r'\d{1,3}(?:,\d{3})+', s)
    if m:
        return float(re.sub(r'[^\d]', '', m.group()))

    # 화폐기호 뒤 소수점 숫자 ($150.00 등)
    m = re.search(r'[\$¥€£₩]\s*(\d+(?:\.\d+)?)', s)
    if m:
        return float(m.group(1))

    # 5자리 이상 순수 숫자
    m = re.search(r'\d{5,}', s)
    if m:
        return float(m.group())

    return 0.0

def get_price_vals(inner_text: str) -> list:
    """
    innerText 전체에서 가격 숫자 목록을 순서대로 반환 (중복 제거).
    핵심 필터:
      - 화폐 기호가 있거나, 쉼표숫자/5자리이상 숫자가 포함된 줄만 처리
      - 할인율(20%)·별점(4.5)·리뷰수(140) 등 노이즈 숫자 자동 제거
      - 최소값 1,000 이상만 가격으로 인정 (소수점 달러는 별도 처리)
    """
    lines = [l.strip() for l in inner_text.split('\n') if l.strip()]
    seen = set()
    result = []
    for line in lines:
        # 퍼센트·괄호 제거 후 가격 패턴 존재 여부 확인
        no_noise = re.sub(r'\b\d{1,3}\s*%', '', line)
        no_noise = re.sub(r'\(\s*\d+\s*\)', '', no_noise)
        has_marker     = any(m in line for m in PRICE_MARKERS)
        has_price_pat  = bool(re.search(
            r'\d{1,3}(?:,\d{3})+|\d{5,}|[\$¥€£₩]\s*\d+', no_noise
        ))
        if not (has_marker or has_price_pat):
            continue
        val = parse_price_from_line(line)
        # 최소 1,000 이상 (달러는 100 이상)
        threshold = 100 if any(m in line for m in ['$', '¥', '€', '£']) else 1000
        if val >= threshold and val not in seen:
            seen.add(val)
            result.append(val)
    return result

def fmt_price(val: float) -> str:
    if val == int(val):
        return f"{int(val):,}"
    return f"{val:,.0f}"

def fmt_sale_with_pct(sale_val: float, reg_val: float) -> str:
    """'188,100 (20%)' 형식으로 세일가+할인율 반환"""
    sale_str = fmt_price(sale_val)
    if reg_val > 0 and 0 < sale_val < reg_val:
        pct = round((reg_val - sale_val) / reg_val * 100)
        return f"{sale_str} ({pct}%)"
    return sale_str


# ─────────────────────────────────────────
# 세일가 정밀 판별
#
# [전략]
# 1) strike/del/s 태그 또는 CSS line-through → 취소선 = 확실한 정가
# 2) 클래스명 키워드 (정가+세일가 모두 발견 시만 반환)
# 3) innerText 줄별 파싱 fallback
#    → 가격 줄 1개: 정가만 / 2개: 큰값=정가·작은값=세일가
# ─────────────────────────────────────────
def get_refined_prices(driver, item_element, product_name=""):
    warn = ""
    try:
        # innerText가 .text보다 headless에서 안정적
        try:
            full_text = driver.execute_script(
                "return arguments[0].innerText;", item_element
            ) or ""
        except Exception:
            full_text = item_element.text or ""

        # ── 1단계: 취소선 HTML 태그 ──────────────────────
        for sel in ["strike", "del", "s",
                    "span[style*='line-through']",
                    "p[style*='line-through']"]:
            tags = item_element.find_elements(By.CSS_SELECTOR, sel)
            if not tags:
                continue
            tag = tags[0]
            reg_text = tag.text.strip()
            if not reg_text:
                try:
                    inner = driver.execute_script("return arguments[0].innerHTML;", tag)
                    reg_text = re.sub(r'<[^>]+>', '', inner).strip()
                except Exception:
                    pass
            reg_val = parse_price_from_line(reg_text)
            if reg_val < 100:
                continue
            # 나머지 텍스트에서 정가보다 작은 가격 탐색
            rest_vals = [v for v in get_price_vals(
                full_text.replace(reg_text, "", 1)
            ) if 0 < v < reg_val]
            if rest_vals:
                sale_val = max(rest_vals)
                return fmt_price(reg_val), fmt_sale_with_pct(sale_val, reg_val), ""
            break  # 취소선 있지만 세일가 없음 → 다음 단계로

        # ── 2단계: JS computed style line-through ────────
        for tag in item_element.find_elements(
            By.CSS_SELECTOR, "span, p, em, strong, b"
        )[:20]:
            try:
                td = driver.execute_script(
                    "return window.getComputedStyle(arguments[0]).textDecoration;", tag
                )
                if not (td and "line-through" in td):
                    continue
                reg_val = parse_price_from_line(tag.text.strip())
                if reg_val < 100:
                    continue
                rest_vals = [v for v in get_price_vals(
                    full_text.replace(tag.text.strip(), "", 1)
                ) if 0 < v < reg_val]
                if rest_vals:
                    sale_val = max(rest_vals)
                    return fmt_price(reg_val), fmt_sale_with_pct(sale_val, reg_val), ""
                break
            except Exception:
                continue

        # ── 3단계: 클래스명 키워드 ───────────────────────
        ORIGIN_KW = [
            "consumer", "origin", "original", "origin-price", "originprice",
            "regular", "regular-price", "regularprice",
            "old", "old-price", "oldprice", "before", "before-price",
            "crossed", "retail", "list-price", "listprice",
            "was", "was-price", "msrp", "compare", "compare-at",
            "normal", "normal-price", "normalprice", "price-origin",
        ]
        SALE_KW = [
            "selling", "sale-price", "saleprice", "price-sale",
            "discount", "discounted", "special", "special-price",
            "final", "final-price", "offer", "offer-price",
            "promo", "promo-price", "saving",
        ]
        reg_val_kw = None
        sale_val_kw = None
        for tag in item_element.find_elements(
            By.CSS_SELECTOR, "span, div, p, em, strong, b"
        )[:40]:
            try:
                combined = (
                    (tag.get_attribute("class") or "") + " " +
                    (tag.get_attribute("id") or "")
                ).lower()
                val = parse_price_from_line(tag.text.strip())
                if val < 100:
                    continue
                if reg_val_kw is None and any(kw in combined for kw in ORIGIN_KW):
                    reg_val_kw = val
                elif sale_val_kw is None and any(kw in combined for kw in SALE_KW):
                    sale_val_kw = val
            except Exception:
                continue

        if reg_val_kw is not None and sale_val_kw is not None:
            return fmt_price(reg_val_kw), fmt_sale_with_pct(sale_val_kw, reg_val_kw), ""
        if reg_val_kw is not None:
            rest = [v for v in get_price_vals(full_text) if 0 < v < reg_val_kw]
            if rest:
                return fmt_price(reg_val_kw), fmt_sale_with_pct(max(rest), reg_val_kw), ""
            return fmt_price(reg_val_kw), "-", ""

        # ── 4단계: innerText 줄별 파싱 fallback ──────────
        # 핵심: 할인율(%)/리뷰수/(괄호숫자) 등 노이즈를 걷어낸 순수 가격 목록
        price_vals = get_price_vals(full_text)

        if len(price_vals) >= 2:
            sorted_vals = sorted(price_vals, reverse=True)
            reg_val  = sorted_vals[0]
            sale_val = sorted_vals[1]
            return fmt_price(reg_val), fmt_sale_with_pct(sale_val, reg_val), ""
        elif len(price_vals) == 1:
            return fmt_price(price_vals[0]), "-", ""

    except Exception as e:
        warn = f"가격 추출 예외: {e}"

    warn = warn or "가격 정보를 찾을 수 없음"
    return "정보없음", "-", warn


# ─────────────────────────────────────────
# 스크롤
# ─────────────────────────────────────────
def scroll_to_bottom(driver, log_fn, pause: float = 0.6, max_rounds: int = 40):
    last_height = driver.execute_script("return document.body.scrollHeight")
    for i in range(max_rounds):
        driver.execute_script("window.scrollBy(0, window.innerHeight * 2);")
        time.sleep(pause)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            time.sleep(pause)
            if driver.execute_script("return document.body.scrollHeight") == last_height:
                log_fn(f"  ✅ 스크롤 완료 ({i+1}회, 총 높이 {new_height}px)")
                break
            last_height = driver.execute_script("return document.body.scrollHeight")
        else:
            last_height = new_height
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(0.3)
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(pause)


# ─────────────────────────────────────────
# 이미지 병렬 다운로드
# ─────────────────────────────────────────
REQ_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                  '(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'image/webp,image/apng,image/*,*/*;q=0.8',
    'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8',
}

def download_single_image(args):
    img_url, save_path, thumb_path = args
    try:
        res = requests.get(img_url, timeout=7, headers=REQ_HEADERS)
        if res.status_code != 200:
            return False, f"HTTP {res.status_code}"
        img_pil = PILImage.open(BytesIO(res.content)).convert("RGB")
        img_pil.save(save_path, "JPEG", quality=85)
        img_thumb = img_pil.copy()
        img_thumb.thumbnail((220, 220))
        img_thumb.save(thumb_path, "PNG")
        return True, ""
    except Exception as e:
        return False, str(e)


# ─────────────────────────────────────────
# 로그 렌더링 (고정 높이 스크롤 박스)
# ─────────────────────────────────────────
def render_logs(logs: list, placeholder):
    colored = []
    for line in logs:
        escaped = (line
                   .replace("&", "&amp;")
                   .replace("<", "&lt;")
                   .replace(">", "&gt;"))
        if any(x in line for x in ["❌", "⚠️", "오류", "실패", "경고"]):
            colored.append(f'<span style="color:#ff6b6b">{escaped}</span>')
        elif any(x in line for x in ["✅", "완료", "🎉"]):
            colored.append(f'<span style="color:#69db7c">{escaped}</span>')
        elif line.startswith("  "):
            colored.append(f'<span style="color:#a9b7c6">{escaped}</span>')
        else:
            colored.append(f'<span style="color:#ffd43b">{escaped}</span>')

    lines_html = "<br>".join(colored)
    html = f"""
<div id="log-box" style="
    height:220px;overflow-y:auto;background:#0e1117;color:#d4d4d4;
    font-family:monospace;font-size:12px;padding:10px 14px;
    border-radius:6px;border:1px solid #333;line-height:1.8;
">{lines_html}</div>
<script>
  (function(){{
    var b=document.getElementById('log-box');
    if(b) b.scrollTop=b.scrollHeight;
  }})();
</script>
"""
    placeholder.markdown(html, unsafe_allow_html=True)


# ─────────────────────────────────────────
# 버튼 + 로그 (항상 고정 순서 렌더링)
# ─────────────────────────────────────────
start_btn = st.button("🚀 데이터 수집 시작", type="primary")

log_placeholder = st.empty()
if st.session_state.logs:
    render_logs(st.session_state.logs, log_placeholder)


# ─────────────────────────────────────────
# 수집 실행
# ─────────────────────────────────────────
if start_btn:
    if not target_url.strip():
        st.error("URL을 입력해주세요.")
        st.stop()

    # 새 수집 시작 → 이전 데이터 초기화
    st.session_state.logs = []
    st.session_state.excel_data = None
    st.session_state.zip_data = None
    st.session_state.result_count = None

    IMG_FOLDER = "collected_images"
    if os.path.exists(IMG_FOLDER):
        shutil.rmtree(IMG_FOLDER)
    os.makedirs(IMG_FOLDER)

    def log(msg: str):
        st.session_state.logs.append(msg)
        render_logs(st.session_state.logs, log_placeholder)

    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )

    driver = None
    try:
        # ── STEP 1 ───────────────────────────────────────
        log("🌐 [1단계] 브라우저 실행 및 페이지 접속 중...")
        driver = webdriver.Chrome(options=options)
        driver.set_page_load_timeout(30)
        driver.get(target_url)
        time.sleep(3)
        log(f"  ✅ 페이지 접속 완료: {driver.title or target_url}")

        # ── STEP 2 ───────────────────────────────────────
        log("📜 [2단계] 전체 상품 로드를 위해 페이지 스크롤 시작...")
        scroll_to_bottom(driver, log, pause=0.6, max_rounds=40)

        # ── STEP 3 ───────────────────────────────────────
        log("🔍 [3단계] 페이지에서 상품 요소 탐색 중...")

        SELECTORS = [
            # 카페24 계열 (코닥 등)
            "ul.prdList > li",
            ".prdList li",
            "div.xans-product-listitem",
            # 상품 카드
            "[class*='product-card']", "[class*='ProductCard']", "[class*='product_card']",
            "[class*='item-card']", "[class*='ItemCard']",
            # 그리드/리스트
            "[class*='product-item']", "[class*='productItem']", "[class*='product_item']",
            "[class*='goods-item']", "[class*='goodsItem']", "[class*='item-wrap']",
            # article/li
            "article[class*='product']", "article[class*='item']",
            "li[class*='product']", "li[class*='item']", "li[class*='goods']",
            # 링크 기반
            "a[class*='product']", "a[class*='item-link']",
            # 폴백
            "li",
        ]

        items = []
        used_selector = ""
        for sel in SELECTORS:
            try:
                candidates = driver.find_elements(By.CSS_SELECTOR, sel)
            except Exception:
                continue
            valid = []
            for c in candidates:
                try:
                    if c.size['width'] < 80 or c.size['height'] < 80:
                        continue
                    if c.find_elements(By.TAG_NAME, 'a') and c.find_elements(By.TAG_NAME, 'img'):
                        valid.append(c)
                except Exception:
                    continue
            if len(valid) >= 3:
                items = valid
                used_selector = sel
                break

        log(f"  ✅ 상품 후보 {len(items)}개 발견 (셀렉터: '{used_selector}')")

        if not items:
            log("⚠️ 상품 요소를 찾지 못했습니다.")
            st.warning("상품 정보를 찾을 수 없습니다. URL을 확인해주세요.")
            driver.quit()
            driver = None
            st.stop()

        # ── STEP 4 ───────────────────────────────────────
        log(f"📦 [4단계] {len(items)}개 상품 정보 추출 중...")

        final_results = []
        seen_links = set()
        skip_count = 0
        warn_logs = []

        for idx, item in enumerate(items, start=1):
            item_warn = []
            try:
                # 링크
                try:
                    link_tag = item if item.tag_name == 'a' else item.find_element(By.TAG_NAME, 'a')
                    link = link_tag.get_attribute('href') or ""
                except Exception:
                    skip_count += 1
                    continue

                if not link or link in seen_links or "javascript" in link.lower():
                    skip_count += 1
                    continue

                item_text = item.text.strip()
                if len(item_text) < 5:
                    skip_count += 1
                    continue

                # 상품명
                lines = [l.strip() for l in item_text.split('\n') if l.strip()]
                p_name = lines[0] if lines else "이름없음"
                if len(p_name) < 3 and len(lines) > 1:
                    p_name = lines[1]

                # 가격
                reg_p, sale_p, price_warn = get_refined_prices(driver, item, p_name)
                if price_warn:
                    item_warn.append(f"가격: {price_warn}")

                # 이미지
                img_urls = []
                try:
                    for img in item.find_elements(By.TAG_NAME, "img"):
                        src = (
                            img.get_attribute('data-src')
                            or img.get_attribute('data-lazy-src')
                            or img.get_attribute('data-original')
                            or img.get_attribute('src')
                            or ""
                        )
                        if not src:
                            srcset = img.get_attribute('srcset') or ""
                            if srcset:
                                src = srcset.strip().split()[0]
                        if not src:
                            continue
                        src_lower = src.lower()
                        if any(x in src_lower for x in [
                            'swatch', 'color', 'icon', 'logo', 'banner',
                            'pixel', 'spacer', 'blank', '1x1', 'placeholder', 'loading'
                        ]):
                            continue
                        img_urls.append(urljoin(target_url, src))
                        if len(img_urls) >= 2:
                            break
                except Exception as e:
                    item_warn.append(f"이미지 추출 예외: {e}")

                if not img_urls:
                    item_warn.append("이미지 URL 없음")
                    skip_count += 1
                    if item_warn:
                        warn_logs.append(f"  ⚠️ [{idx}] {p_name[:40]} → {' / '.join(item_warn)}")
                    continue

                if item_warn:
                    warn_logs.append(f"  ⚠️ [{idx}] {p_name[:40]} → {' / '.join(item_warn)}")

                final_results.append({
                    "제품명": p_name[:80],
                    "정가": reg_p,
                    "세일가": sale_p,
                    "링크": link,
                    "이미지들": img_urls,
                })
                seen_links.add(link)

                if idx % 20 == 0:
                    log(f"  📦 {idx}/{len(items)} 처리 중... (수집됨: {len(final_results)}개)")

            except Exception as e:
                skip_count += 1
                warn_logs.append(f"  ⚠️ [{idx}] 상품 파싱 예외: {e}")
                continue

        log(f"  ✅ 추출 완료 → 유효 {len(final_results)}개 / 스킵 {skip_count}개")

        if warn_logs:
            log(f"  ─── 개별 상품 경고 ({len(warn_logs)}건, 최대 30건 표시) ───")
            for w in warn_logs[:30]:
                log(w)

        if not final_results:
            log("⚠️ 유효한 상품이 없어 수집을 종료합니다.")
            st.warning("상품 정보를 추출하지 못했습니다.")
            driver.quit()
            driver = None
            st.stop()

        driver.quit()
        driver = None
        log("🌐 브라우저 종료")

        # ── STEP 5 ───────────────────────────────────────
        log(f"🖼️ [5단계] 이미지 병렬 다운로드 시작 ({len(final_results)}개 상품)...")

        download_tasks = []
        for i, data in enumerate(final_results, start=1):
            for j, img_url in enumerate(data["이미지들"]):
                save_path  = os.path.join(IMG_FOLDER, f"{i}_{j+1}.jpg")
                thumb_path = os.path.join(IMG_FOLDER, f"t_{i+1}_{j}.png")
                download_tasks.append((img_url, save_path, thumb_path, i, j))

        img_results = {}
        completed_dl = 0
        dl_fail_logs = []

        with ThreadPoolExecutor(max_workers=12) as executor:
            future_map = {
                executor.submit(download_single_image, (url, sp, tp)): (ii, jj, tp, url)
                for url, sp, tp, ii, jj in download_tasks
            }
            for future in as_completed(future_map):
                ii, jj, tp, url = future_map[future]
                ok, err_msg = future.result()
                img_results[(ii, jj)] = tp if ok else None
                if not ok:
                    dl_fail_logs.append(
                        f"  ⚠️ 이미지 다운 실패 [상품#{ii} img{jj+1}]: {err_msg} | {url[:55]}..."
                    )
                completed_dl += 1
                if completed_dl % 20 == 0:
                    log(f"  🖼️ 이미지 {completed_dl}/{len(download_tasks)} 다운로드 완료...")

        log(f"  ✅ 이미지 다운로드 완료 ({completed_dl}개 처리, 실패 {len(dl_fail_logs)}건)")
        for w in dl_fail_logs[:20]:
            log(w)

        # ── STEP 6 ───────────────────────────────────────
        log("📊 [6단계] 엑셀 파일 생성 중...")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "상품목록"

        headers = ["번호", "제품명", "이미지1", "이미지2", "정가", "세일가", "상세링크"]
        ws.append(headers)

        # 헤더: 흰 배경 + 검정 볼드
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="000000")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(fill_type=None)

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 28
        ws.column_dimensions['D'].width = 28
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 26   # 세일가+할인율 표시용
        ws.column_dimensions['G'].width = 14

        for i, data in enumerate(final_results, start=1):
            row_idx = i + 1
            ws.row_dimensions[row_idx].height = 165

            ws.cell(row=row_idx, column=1, value=i).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row_idx, column=2, value=data["제품명"]).alignment = Alignment(wrap_text=True, vertical='center')
            ws.cell(row=row_idx, column=5, value=data["정가"]).alignment = Alignment(horizontal='center', vertical='center')

            # 세일가
            s_cell = ws.cell(row=row_idx, column=6)
            s_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if data["세일가"] != "-":
                s_cell.value = data["세일가"]   # 예: '188,100 (20%)'
                s_cell.font = Font(color="CC0000", bold=True, size=11)
            else:
                s_cell.value = "-"

            # 링크
            link_cell = ws.cell(row=row_idx, column=7, value="링크")
            link_cell.hyperlink = data["링크"]
            link_cell.font = Font(color="0563C1", underline="single")
            link_cell.alignment = Alignment(horizontal='center', vertical='center')

            # 이미지 삽입
            for j in range(2):
                tp = img_results.get((i, j))
                if tp and os.path.exists(tp):
                    col_letter = 'C' if j == 0 else 'D'
                    try:
                        ws.add_image(XLImage(tp), f"{col_letter}{row_idx}")
                    except Exception as e:
                        log(f"  ⚠️ 이미지 삽입 실패 [행{row_idx}, img{j+1}]: {e}")

        excel_io = BytesIO()
        wb.save(excel_io)
        st.session_state.excel_data = excel_io.getvalue()
        log("  ✅ 엑셀 파일 생성 완료")

        # ── STEP 7 ───────────────────────────────────────
        log("🗜️ [7단계] 이미지 ZIP 압축 중...")

        zip_io = BytesIO()
        with zipfile.ZipFile(zip_io, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for root, _, files in os.walk(IMG_FOLDER):
                for f in files:
                    if not f.startswith("t_"):
                        zf.write(os.path.join(root, f), f)
        st.session_state.zip_data = zip_io.getvalue()
        st.session_state.result_count = len(final_results)
        log("  ✅ ZIP 압축 완료")

        shutil.rmtree(IMG_FOLDER)
        log(f"🎉 모든 작업 완료! 총 {len(final_results)}개 상품 수집됨.")

    except Exception as e:
        log(f"❌ 치명적 오류 발생: {e}")
        st.error(f"오류: {e}")
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass
        if os.path.exists(IMG_FOLDER):
            shutil.rmtree(IMG_FOLDER)

    st.rerun()


# ─────────────────────────────────────────
# 결과 다운로드 버튼
# ─────────────────────────────────────────
if st.session_state.excel_data:
    st.divider()
    st.success(f"✅ {st.session_state.result_count}개 상품 수집 완료!")
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    c1, c2 = st.columns(2)
    with c1:
        st.download_button(
            label="📥 엑셀 다운로드",
            data=st.session_state.excel_data,
            file_name=f"products_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with c2:
        st.download_button(
            label="🖼️ 이미지(ZIP) 다운로드",
            data=st.session_state.zip_data,
            file_name=f"images_{ts}.zip",
            mime="application/zip",
            use_container_width=True,
        )